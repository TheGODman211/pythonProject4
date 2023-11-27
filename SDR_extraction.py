
import openpyxl
import glob
import os
from openpyxl import load_workbook

import types
import os, re


# path18 = r"D:\excel\Bew folder\2018 Annual SDR\Life 2018"
# path19 = r"D:\excel\Bew folder\Annual SDR 2019\LIfe"
# path20 = r"D:\excel\Bew folder\Annual SDR 2020\Life"
# path22 = r"D:\excel\Bew folder\2022\Annual 2022\LIFE COMPANIES"
# path21 = r"D:\excel\Bew folder\Annual\Annual 2021 Life"

path18 = r"D:\excel\Bew folder\2018 Annual SDR\NON-LIFE 2018"
path19 = r"D:\excel\Bew folder\Annual SDR 2019\Non-life"
path20 = r"D:\excel\Bew folder\Annual SDR 2020\Non-life 2020"
path22 = r"D:\excel\Bew folder\2022\Annual 2022\NON LIFE COMPANIES"
path21 = r"D:\excel\Bew folder\Annual\Annual 2021 Non-Life"
excel_file = glob.glob(r"D:\excel\Bew folder\2023\Q2\Non life\[!~]*.xlsx", recursive=True)


# ---- FUNCTION TO SEARCH FOR FILES IN SPECIFIED PATH AND SAVE TO A LIST
excel_file18 = glob.glob(os.path.join(path18, '[!~]*.xlsx'))
excel_file19 = glob.glob(os.path.join(path19, '[!~]*.xlsx'))
excel_file20 = glob.glob(os.path.join(path20, '[!~]*.xlsx'))
excel_file21 = glob.glob(os.path.join(path21, '[!~]*.xlsx'))
excel_file22 = glob.glob(os.path.join(path22, '[!~]*.xlsx'))

print(len(excel_file21))
gp ,np,ni,gbp,nbp,me,ce,ur,ini,oi,ci,pat,cb,ina,rec,ppe,ta,tp,pay,name,ul,gl,tl,cl,nle,dp,anu,tpd,oap,gsp,ol,suy,ash,pt,pt1,pt2,pt3,pt4,pt5,pt6,pt7 =[],[], [],[],[],[],[],[],[],[], [],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]



for i, file in enumerate(excel_file, start=0):
    wb=load_workbook(file, data_only=True)
    print(file)
    wb.active=wb['SDR8ii']
    np.append(wb.active.cell(row=9, column=4).value)
    ni.append(wb.active.cell(row=10, column=4).value)
    gbp.append(wb.active.cell(row=11, column=4).value)
    nbp.append(wb.active.cell(row=12, column=4).value)
    me.append(wb.active.cell(row=13, column=4).value)
    ce.append(wb.active.cell(row=14, column=4).value)
    ur.append(wb.active.cell(row=15, column=4).value)
    ini.append(wb.active.cell(row=20, column=4).value)
    oi.append(wb.active.cell(row=21, column=4).value)
    ci.append(wb.active.cell(row=22, column=4).value)
    ul.append(wb.active.cell(row=23, column=4).value)
    gl.append(wb.active.cell(row=24, column=4).value)
    tl.append(wb.active.cell(row=25, column=4).value)
    cl.append(wb.active.cell(row=26, column=4).value)
    nle.append(wb.active.cell(row=27, column=4).value)
    dp.append(wb.active.cell(row=28, column=4).value)
    anu.append(wb.active.cell(row=29, column=4).value)
    tpd.append(wb.active.cell(row=30, column=4).value)
    oap.append(wb.active.cell(row=31, column=4).value)
    gsp.append(wb.active.cell(row=32, column=4).value)
    ol.append(wb.active.cell(row=33, column=4).value)
    suy.append(wb.active.cell(row=34, column=4).value)
    ash.append(wb.active.cell(row=35, column=4).value)
    # gra.append(wb.active.cell(row=31, column=5).value)
    # kum.append(wb.active.cell(row=31, column=5).value)
    # iop.append(wb.active.cell(row=31, column=5).value)

    # wb.active = wb['SDR4i']
    # ni.append(wb.active.cell(row=11, column=3).value)




    wb.active = wb['SDR4i']
    pt.append(wb.active.cell(row=14, column=6).value)
    pt1.append(wb.active.cell(row=15, column=6).value)
    pt2.append(wb.active.cell(row=16, column=6).value)
    pt3.append(wb.active.cell(row=17, column=6).value)
    pt4.append(wb.active.cell(row=18, column=6).value)
    pt5.append(wb.active.cell(row=19, column=6).value)
    pt6.append(wb.active.cell(row=20, column=6).value)
    pt7.append(wb.active.cell(row=21, column=6).value)
    # oap.append(wb.active.cell(row=19, column=5).value)
    # print(ul)

    wb.active = wb['SDR1']
    name.append(wb.active.cell(row=1, column=2).value)
    np.append(wb.active.cell(row=95, column=3).value)



    print("done")
    # print(np[-1])


ws=openpyxl.Workbook()
for index, value in enumerate(ni, start=2):
    ws.active.cell(row=index, column=2).value = np[index - 2]
    ws.active.cell(row=index, column=3).value = ni[index - 2]
    ws.active.cell(row=index, column=1).value = name[index - 2]
    ws.active.cell(row=index, column=13).value = ul[index - 2]
    ws.active.cell(row=index, column=4).value = gbp[index - 2]
    ws.active.cell(row=index, column=5).value = nbp[index - 2]
    ws.active.cell(row=index, column=13).value = gl[index - 2]
    ws.active.cell(row=index, column=14).value = tl[index - 2]
    ws.active.cell(row=index, column=16).value = cl[index - 2]
    ws.active.cell(row=index, column=17).value = nle[index - 2]
    ws.active.cell(row=index, column=18).value = dp[index - 2]
    ws.active.cell(row=index, column=19).value = anu[index - 2]
    ws.active.cell(row=index, column=20).value = tpd[index - 2]
    ws.active.cell(row=index, column=21).value = oap[index - 2]
    ws.active.cell(row=index, column=22).value = gsp[index - 2]
    ws.active.cell(row=index, column=23).value = ol[index - 2]
    ws.active.cell(row=index, column=24).value = suy[index - 2]
    ws.active.cell(row=index, column=25).value = ash[index - 2]
    ws.active.cell(row=index, column=26).value = pt[index - 2]
    ws.active.cell(row=index, column=27).value = pt1[index - 2]
    ws.active.cell(row=index, column=28).value = pt2[index - 2]
    ws.active.cell(row=index, column=29).value = pt3[index - 2]
    ws.active.cell(row=index, column=30).value = pt4[index - 2]
    ws.active.cell(row=index, column=31).value = pt5[index - 2]
    ws.active.cell(row=index, column=32).value = pt6[index - 2]
    ws.active.cell(row=index, column=33).value = pt7[index - 2]

    ws.active.cell(row=index, column=6).value = me[index - 2]
    ws.active.cell(row=index, column=7).value = ce[index - 2]

    ws.active.cell(row=index, column=8).value = ur[index - 2]
    ws.active.cell(row=index, column=9).value = ini[index - 2]
    ws.active.cell(row=index, column=10).value = oi[index - 2]
    ws.active.cell(row=index, column=11).value = ci[index - 2]
#
#
ws.active.cell(row=1, column=1).value ="Name"
ws.active.cell(row=1, column=2).value ="CAR"
ws.active.cell(row=1, column=3).value ="MCR"
ws.active.cell(row=1, column=4).value ="RE"
# ws.active.cell(row=1, column=5).value ="UR"
# ws.active.cell(row=1, column=2).value ="CLAIMS PAID"
# ws.active.cell(row=1, column=27).value ="MGT"
# ws.active.cell(row=1, column=28).value ="TOTAL EXPENSE"
# ws.active.cell(row=1, column=29).value ="PAT"

ws.save(r'D:\excel\lob23.xlsx')
