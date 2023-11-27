from openpyxl import load_workbook
import openpyxl
import glob
#import pyexcel as p
excel_file = glob.glob(r"D:\excel\Bew folder\2023\Q2\Life\[!~]*.xlsx", recursive=True)
#excel_file = glob.glob("D:\excel\Bew folder\2023\NOn Life\*.xlsx")
print(excel_file)
new = []
for file in excel_file:
    new.append(file[:-4])
gp,tni,pbp ,np,ni,gbp,nbp,me,ce,ur,ini,oi,ci,pat,cb,ina,rec,ppe,ta,tp,pay,name,ul,gl,tl,cl,nle,dp,anu,tpd,oap,gsp,ol = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
for i,file in enumerate(excel_file, start=0):
    # if file[-3:] == 'xls':
    #     p.save_book_as(filename = file, dest_filename = new[i] )

    wb = load_workbook(file,  data_only=True)
    wb.active = wb["SDR3"]
    ol.append(wb.active.cell(row =12, column=4).value)
    gsp.append(wb.active.cell(row =16, column=4).value)
    if type(ol[i]) == type(None):
        ol[i] = 0

    if type(gsp[i]) == type(None):
        gsp[i] = 0
    print(gsp)
    print(ol)
    gp.append(ol[i]+gsp[i])
    np.append(wb.active.cell(row =15, column=4).value)
    ni.append(wb.active.cell(row =21, column=4).value)
    tni.append(wb.active.cell(row=18, column=4).value)
    pbp.append(wb.active.cell(row=20, column=4).value)
    gbp.append(wb.active.cell(row =24, column=4).value)
    nbp.append(wb.active.cell(row =26, column=4).value)
    me.append(wb.active.cell(row =30, column=4).value)
    ce.append(wb.active.cell(row =27, column=4).value)
    ur.append(wb.active.cell(row =35, column=4).value)
    ini.append(wb.active.cell(row =44, column=4).value)
    oi.append(wb.active.cell(row =49, column=4).value)
    ci.append(wb.active.cell(row =33, column=4).value)
    pat.append(wb.active.cell(row =53, column=4).value)


    wb.active = wb['SDR2']
    name.append(wb.active.cell(row =1, column=2).value)
    cb.append(wb.active.cell(row =9, column=3).value)
    ina.append(wb.active.cell(row=26, column=3).value)
    rec.append(wb.active.cell(row=41, column=3).value)
    ppe.append(wb.active.cell(row=51, column=3).value)
    ta.append(wb.active.cell(row=65, column=3).value)

    wb.active = wb['SDR2i']
    tp.append(wb.active.cell(row=14, column=3).value)
    pay.append(wb.active.cell(row=23, column=3).value)

    wb.active =wb['SDR8i']
    ul.append(wb.active.cell(row=11, column=5).value)
    gl.append(wb.active.cell(row=12, column=5).value)
    tl.append(wb.active.cell(row=13, column=5).value)
    cl.append(wb.active.cell(row=14, column=5).value)
    nle.append(wb.active.cell(row=15, column=5).value)
    dp.append(wb.active.cell(row=16, column=5).value)
    anu.append(wb.active.cell(row=17, column=5).value)
    tpd.append(wb.active.cell(row=18, column=5).value)
    oap.append(wb.active.cell(row=19, column=5).value)

ws = openpyxl.Workbook()
#ws = load_workbook('D:\excel\LifeQ1.xlsx',  data_only=True)
print(gp)

print(excel_file)
for index, value in enumerate(gp, start=2):
    ws.active.cell(row=index,column=1).value = gp[index-2]
    ws.active.cell(row=index, column=2).value = np[index-2]
    ws.active.cell(row=index, column=3).value = ni[index - 2]
    ws.active.cell(row=index, column=4).value = gbp[index - 2]
    ws.active.cell(row=index, column=5).value = nbp[index - 2]
    ws.active.cell(row=index, column=6).value = me[index - 2]
    ws.active.cell(row=index, column=7).value = ce[index - 2]
    ws.active.cell(row=index, column=9).value = ur[index - 2]
    ws.active.cell(row=index, column=10).value = ini[index - 2]
    ws.active.cell(row=index, column=11).value = oi[index - 2]
    ws.active.cell(row=index, column=8).value = ci[index - 2]
    ws.active.cell(row=index, column=12).value = pat[index - 2]
    ws.active.cell(row=index, column=13).value = cb[index - 2]
    ws.active.cell(row=index, column=14).value = ina[index - 2]
    ws.active.cell(row=index, column=15).value = rec[index - 2]
    ws.active.cell(row=index, column=16).value = ppe[index - 2]
    ws.active.cell(row=index, column=17).value = ta[index - 2]
    ws.active.cell(row=index, column=18).value = tp[index - 2]
    ws.active.cell(row=index, column=19).value = pay[index - 2]
    ws.active.cell(row=index, column=20).value = name[index - 2]
    ws.active.cell(row=index, column=21).value = ul[index - 2]
    ws.active.cell(row=index, column=22).value = gl[index - 2]
    ws.active.cell(row=index, column=23).value = tl[index - 2]
    ws.active.cell(row=index, column=24).value = cl[index - 2]
    ws.active.cell(row=index, column=25).value = nle[index - 2]
    ws.active.cell(row=index, column=26).value = dp[index - 2]
    ws.active.cell(row=index, column=27).value = anu[index - 2]
    ws.active.cell(row=index, column=28).value = tpd[index - 2]
    ws.active.cell(row=index, column=29).value = oap[index - 2]
    ws.active.cell(row=index, column=30).value = tni[index - 2]
    ws.active.cell(row=index, column=31).value = pbp[index - 2]





    print(index)

ws.active.cell(row=1, column=1).value ="Gross premium"
ws.active.cell(row=1, column=2).value ="Net premium"
ws.active.cell(row=1, column=3).value ="Net income"
ws.active.cell(row=1, column=4).value ="Gross benefit payed"
ws.active.cell(row=1, column=5).value ="Net benefit payed"
ws.active.cell(row=1, column=6).value ="mgt expense"
ws.active.cell(row=1, column=7).value ="Comission expense"
ws.active.cell(row=1, column=9).value ="Underwriting Results"
ws.active.cell(row=1, column=10).value ="Investment Income"
ws.active.cell(row=1, column=11).value ="Other Income"
ws.active.cell(row=1, column=8).value ="Commission Income"
ws.active.cell(row=1, column=12).value ="PAT"
ws.active.cell(row=1, column=13).value ="Cash Balance"
ws.active.cell(row=1, column=14).value ="Inv Assest"
ws.active.cell(row=1, column=15).value ="Receivables"
ws.active.cell(row=1, column=16).value ="PPEs"
ws.active.cell(row=1, column=17).value ="Total Asset"
ws.active.cell(row=1, column=18).value ="Technical Provision"
ws.active.cell(row=1, column=19).value ="Payables"
ws.active.cell(row=1, column=21).value ="Universal"
ws.active.cell(row=1, column=22).value ="Group"
ws.active.cell(row=1, column=23).value ="Term"
ws.active.cell(row=1, column=24).value ="Credit"
ws.active.cell(row=1, column=25).value ="Whole Life"
ws.active.cell(row=1, column=26).value ="Dread disease"
ws.active.cell(row=1, column=27).value ="Annuities"
ws.active.cell(row=1, column=28).value ="Total And Permanent disability"
ws.active.cell(row=1, column=29).value ="Other Approved Product"
ws.active.cell(row=1, column=30).value ="Total Net Inflow"
ws.active.cell(row=1, column=31).value ="Change in policyholder benefit"



ws.save('D:\excel\gpl.xlsx')