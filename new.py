# import os
# from openpyxl import load_workbook
# from math import trunc
#
# folder_path = r"D:\excel\Sign in\raw2"  # Replace with the path to your folder
# output_path = r"D:\excel\Sign in\raw2"
#
# # Iterate through all files in the folder
# for filename in os.listdir(folder_path):
#     if filename.endswith('.xlsx'):
#         file_path = os.path.join(folder_path, filename)
#
#         # Load the workbook
#         wb = load_workbook(file_path)
#         ws = wb.active
#
#         # Get the number of rows in column B
#         num_rows = ws.max_row
#
#         # Create a set to store unique values in column I
#         unique_values = set()
#
#         # Iterate through the rows and perform the calculations
#         for row in range(2, num_rows + 1):
#             b_value = ws[f'B{row}'].value
#             i_value = trunc(b_value)
#             j_value = b_value - i_value
#
#             ws[f'I{row}'] = i_value
#             ws[f'J{row}'] = j_value
#
#             unique_values.add(i_value)
#
#         # Write the unique values to column K
#         for idx, value in enumerate(unique_values, start=2):
#             ws[f'K{idx}'] = value
#
#         # Copy values from I, J, K to L, M, N
#         for row in range(2, num_rows + 1):
#             ws[f'L{row}'] = ws[f'I{row}'].value
#             ws[f'M{row}'] = ws[f'J{row}'].value
#             ws[f'N{row}'] = ws[f'K{row}'].value
#
#         # Copy values from L, M, N to E, F, H
#         for row in range(2, num_rows + 1):
#             ws[f'E{row}'] = ws[f'L{row}'].value
#             ws[f'F{row}'] = ws[f'M{row}'].value
#             ws[f'H{row}'] = ws[f'N{row}'].value
#
#         # Save the changes to the file
#         wb.save(os.path.join(output_path, filename[:-5] + "_fot.xlsx"))
#
# print('Processing complete!')





#ALT 2


import os
from openpyxl import load_workbook
from pynput.keyboard import Key, Controller as KeyboardController
import time

folder_path = r"D:\excel\Sign in\raw2\jan"  # Replace with the path to your folder
our =r"D:\excel\Sign in\raw2\jan"
keyboard = KeyboardController()

# Iterate through all files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)

        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb.active
        ws.cell(row=3, column = 9,value = '=trunc(b3,0')

        # # Get the number of rows in column B
        # num_rows = ws.max_row
        #
        # # Open the file with Excel (update the path to Excel if needed)
        # os.system(f'start excel "{file_path}"')
        #
        # #Wait for Excel to open
        # time.sleep(3)
        #
        # # Type the formulas for columns I and J
        # keyboard.press(Key.ctrl)
        # keyboard.press('g')
        # keyboard.release(Key.ctrl)
        # keyboard.type('I2')
        # keyboard.press(Key.enter)
        # keyboard.release(Key.enter)
        # time.sleep(1)
        # keyboard.type('=TRUNC(B2,0)')
        # time.sleep(1)
        # keyboard.press(Key.enter)
        # keyboard.release(Key.enter)
        #
        # keyboard.press(Key.ctrl)
        # keyboard.press('g')
        # keyboard.release(Key.ctrl)
        # keyboard.type('J2')
        # keyboard.press(Key.enter)
        # keyboard.release(Key.enter)
        # time.sleep(1)
        # keyboard.type('=B2-I2')
        # print('B2-I2')
        # time.sleep(1)
        #
        # keyboard.press(Key.enter)
        # keyboard.release(Key.enter)
        # print('ai')
        #
        # # Save and close the workbook to update the formulas
        # keyboard.press(Key.ctrl)
        # keyboard.press('s')
        # keyboard.release(Key.ctrl)
        # keyboard.release('s')
        # print('saved')
        # time.sleep(1)
        # print('start close')
        # keyboard.press(Key.alt)
        # keyboard.press(Key.f4)
        # keyboard.release(Key.alt)
        # keyboard.release(Key.f4)
        # print("end close")
        # time.sleep(1)






        # # Reopen the workbook with openpyxl
        # wb = load_workbook(file_path)
        # ws = wb.active
        #
        # # Autofill the formulas for columns I and J
        # ws['I2'].auto_fill(ws[f'I3:I{num_rows}'], 'Copy')
        # ws['J2'].auto_fill(ws[f'J3:J{num_rows}'], 'Copy')
        #
        # # Collect unique values in column I
        # unique_values = set(ws[f'I{row}'].value for row in range(2, num_rows + 1))
        #
        # # Write unique values to column K
        # for idx, value in enumerate(unique_values, start=2):
        #     ws[f'K{idx}'] = value
        #
        # # Copy values from I, J, K to L, M, N
        # for row in range(2, num_rows + 1):
        #     ws[f'L{row}'] = ws[f'I{row}'].value
        #     ws[f'M{row}'] = ws[f'J{row}'].value
        #     ws[f'N{row}'] = ws[f'K{row}'].value
        #
        # # Copy values from L, M, N to E, F, H
        # for row in range(2, num_rows + 1):
        #     ws[f'E{row}'] = ws[f'L{row}'].value
        #     ws[f'F{row}'] = ws[f'M{row}'].value
        #     ws[f'H{row}'] = ws[f'N{row}'].value

        # Save the changes to the file
        wb.save(os.path.join(our, filename[:-5] + "_fot.xlsx"))

print('Processing complete!')
