# import subprocess
#
# username = "https://www.instagram.com/veetnigeria/"
# output_directory = r"D:\excel"
# # Download all media
# subprocess.call(f"instagram-scraper {username} --destination {output_directory} --media-metadata --maximum 0 --media-types image,video --profile-metadata", shell=True)



from openpyxl import load_workbook
import openpyxl
import glob

import types
import os, re

path = r"D:\excel\Bew folder\Annual 2022\REINSURER"
path1 = r"D:\excel\Bew folder\Annual 2022\NON LIFE COMPANIES"

# ---- FUNCTION TO SEARCH FOR FILES IN SPECIFIED PATH AND SAVE TO A LIST
excel_file = glob.glob(os.path.join(path, '[!~]*.xlsx'))
nonlife_file = glob.glob(os.path.join(path1, '*.xlsx'))
# excel_file = glob.glob("D:\excel\Q4 2022\Life\*.xlsx", recursive=True) ------this is another to search for the files
print(excel_file)
new = []
current_gross_written_premium,prior_gross_written_premium, current_net_written_premium, prior_net_written_premium, current_equity, prior_equity, outstanding_claims, IBNR, net_earned_premium, cash, investment_income, investment_assets, current_pat, total_expense, net_claims_incurred, total_assets, total_tech_provisions, short_term_liabilities= [],[], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []
gp, np, ni, gbp, nbp, me, ce, ur, ini, oi, ci, pat, cb, ina, rec, ppe, ta, tp, pay, name, ul, gl, tl, cl, nle, dp, anu, tpd, oap = [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []
gog, bog, stat_depo, term_depo, other_depo, money_market=[],[],[],[],[],[]

for i, file in enumerate(excel_file):
    print(file)
    # if file[-3:] == 'xls':
    #     p.save_book_as(filename = file, dest_filename = new[i] )

    wb = load_workbook(file, data_only=True)  # load the whole SDR into python and save to a variable 'wb'
    g_sheet = wb.sheetnames
    # for o in g_sheet:
    #     print(o)
    print(g_sheet[1])
    wb.active = wb[g_sheet[1]]  # select the particular SDR sheet you want to work with
    # select cell value you want and add it to one of the list so can work with it
    name.append(wb.active.cell(row=1, column=2).value)
    # ni.append(wb.active.cell(row =14, column=2).value)     # all these lines of code have been commented out for now
    # gbp.append(wb.active.cell(row =15, column=2).value)
    # nbp.append(wb.active.cell(row =16, column=2).value)
    # ur.append(wb.active.cell(row=17, column=2).value)
    # ini.append(wb.active.cell(row=18, column=2).value)
    # me.append(wb.active.cell(row =20, column=2).value)
    # ce.append(wb.active.cell(row =38, column=2).value)
    print(name)
    wb.active = wb[g_sheet[2]]
    cash.append(wb.active.cell(row=11, column=5).value)
    investment_assets.append(wb.active.cell(row=28, column=5).value)
    total_assets.append(wb.active.cell(row=60, column=5).value)
    gog.append(wb.active.cell(row=14, column=5).value)
    bog.append(wb.active.cell(row=15, column=5).value)
    stat_depo.append(wb.active.cell(row=16, column=5).value)
    term_depo.append(wb.active.cell(row=17, column=5).value)
    other_depo.append(wb.active.cell(row=18, column=5).value)
    money_market.append(wb.active.cell(row=23, column=5).value)
    print(cash)
    wb.active = wb[g_sheet[3]]
    current_equity.append(wb.active.cell(row=58, column=5).value)
    prior_equity.append(wb.active.cell(row=58, column=6).value)
    outstanding_claims.append(wb.active.cell(row=14, column=5).value)
    IBNR.append(wb.active.cell(row=15, column=5).value)
    total_tech_provisions.append(wb.active.cell(row=20, column=5).value)
    short_term_liabilities.append(wb.active.cell(row=30, column=5).value)

    wb.active = wb[g_sheet[6]]
    current_gross_written_premium.append(wb.active.cell(row=13, column=6).value)
    prior_gross_written_premium.append(wb.active.cell(row=13, column=7).value)
    current_net_written_premium.append(wb.active.cell(row=16, column=6).value)
    prior_net_written_premium.append(wb.active.cell(row=16, column=7).value)
    net_earned_premium.append(wb.active.cell(row=20, column=6).value)
    investment_income.append(wb.active.cell(row=43, column=6).value)
    current_pat.append(wb.active.cell(row=52, column=6).value)
    total_expense.append(wb.active.cell(row=29, column=6).value)
    net_claims_incurred.append(wb.active.cell(row=23, column=6).value)

    # wb.active = wb[g_sheet[2]]
    # oi.append(wb.active.cell(row=65, column=3).value)

# for q, file1 in enumerate(nonlife_file):
#     wa = load_workbook(file1, data_only=True)  # load the whole SDR into python and save to a variable 'wb'
#     g_sheet1 = wa.sheetnames
#
#     wa.active = wa[g_sheet1[1]]
#     ni.append(wa.active.cell(row=1, column=2).value)
#
#     wa.active = wa[g_sheet1[14]]
#     gbp.append(wa.active.cell(row=19, column=18).value)

ws = openpyxl.Workbook()
a = 2

# ----WRITE VALUES INTO THE NEW WORKBOOK-----
sheet1 = ws.active
sheet1.title = "Ratios"
# for index, value in enumerate(current_gross_written_premium, start=a):
#     ws.active.cell(row=index, column=1).value = name[index - a]
#     # ws.active.cell(row=index, column=2).value = oi[index-a]
#     ws.active.cell(row=index, column=2).value = gp[index - a]

    #create a new worksheet in ws and write the non life data in it
sheet2 = ws.create_sheet(title="NonLife")
ws.active = sheet2
for index, value in enumerate(current_gross_written_premium, start= a):
    ws.active.cell(row=index, column=1).value = name[index - a]

    # ul is CHANGE IN GROSS WRITTEN PREMIUM
    ul.append((current_gross_written_premium[index-a]-prior_gross_written_premium[index-a])/prior_gross_written_premium[index-a])
    ws.active.cell(row=index, column=2).value = ul[index - a]

    # np is CHANGE IN NET WRITTEN PRERMIUM
    np.append((current_net_written_premium[index-a]-prior_net_written_premium[index-a])/prior_net_written_premium[index-a])
    ws.active.cell(row=index, column=3).value = np[index - a]

    # gbp is NET INSURANCE RISK RATIO
    gbp.append(current_net_written_premium[index-a]/current_equity[index-a])
    ws.active.cell(row=index, column=4).value = gbp[index - a]

    # nbp is GROSS INSURANCE RISK RATIO
    nbp.append(current_gross_written_premium[index-a]/current_equity[index-a])
    ws.active.cell(row=index, column=5).value = nbp[index - a]

    # me is CHANGE IN CAPITAL & SURPLUS
    me.append((current_equity[index-a]-prior_equity[index-a])/prior_equity[index-a])
    ws.active.cell(row=index, column=6).value = me[index - a]

    # ce is CLAIMS RESERVE RATIO
    ce.append((outstanding_claims[index-a]+IBNR[index-a])/net_earned_premium[index-a])
    ws.active.cell(row=index, column=7).value = ce[index - a]

    # ur is RETENTION RATIO
    ur.append(current_net_written_premium[index-a]/current_gross_written_premium[index-a])
    ws.active.cell(row=index, column=8).value = ur[index - a]

    # ini is INVESTMENT YIELD
    ini.append(investment_income[index-a]/(investment_assets[index-a]+cash[index-a]))
    ws.active.cell(row=index, column=9).value = ini[index - a]

    # oi is RETURN ON ASSETS
    oi.append(current_pat[index-a]/total_assets[index-a])
    ws.active.cell(row=index, column=10).value = oi[index - a]

    # ci is EXPENSE RATIO
    ci.append(total_expense[index-a]/net_earned_premium[index-a])
    ws.active.cell(row=index, column=11).value = ci[index - a]

    # pat is LIQUIDITY RATIO
    pat.append((short_term_liabilities[index-a]+total_tech_provisions[index-a])/(gog[index-a]+bog[index-a]+stat_depo[index-a]+term_depo[index-a]+
                                              other_depo[index-a]+money_market[index-a]))
    ws.active.cell(row=index, column=12).value = pat[index - a]

    # cb is TECHNICAL RESERVE COVER
    cb.append(total_tech_provisions[index-a]/(gog[index-a]+bog[index-a]+stat_depo[index-a]+term_depo[index-a]+
                                              other_depo[index-a]+money_market[index-a]))
    ws.active.cell(row=index, column=13).value = cb[index - a]

    #ina is RETURN ON EQUITY
    ina.append(current_pat[index-a]/current_equity[index-a])
    ws.active.cell(row=index, column=14).value = ina[index - a]

    # rec is COMBINED RATIO
    rec.append((net_claims_incurred[index-a]+total_expense[index-a])/net_earned_premium[index-a])
    ws.active.cell(row=index, column=15).value = rec[index - a]

    # ppe is CLAIMS RATIO
    ppe.append(net_claims_incurred[index-a]/net_earned_premium[index-a])
    ws.active.cell(row=index, column=16).value = ppe[index - a]

    # ta is PROPORTION OF INVESTMENT
    ta.append((investment_assets[index-a]+cash[index-a])/total_assets[index-a])
    ws.active.cell(row=index, column=17).value = ta[index - a]

    #to

    #pay

    #ws.active.cell(row=index, column=3).value = gbp[index - 2]
    # ws.active.cell(row=index, column=4).value = gbp[index - 2]
    # ws.active.cell(row=index, column=5).value = nbp[index - 2]
    # ws.active.cell(row=index, column=6).value = ur[index - 2]
    # ws.active.cell(row=index, column=7).value = ini[index - 2]
    # ws.active.cell(row=index, column=8).value = me[index - 2]
    # ws.active.cell(row=index, column=9).value = ce[index - 2]
    # ws.active.cell(row=index, column=10).value = oi[index - 2]
    # print(name[index - 2])
    #
    # if type(gp[index - a]) is types.NoneType:
    #     gp[index - a] = 0
    # if type(gbp[index - a]) is types.NoneType:
    #    gbp[index - a] = 0
    # if type(cb[index - a]) is types.NoneType:
    #     cb[index - a] = 0
    # if type(ina[index - a]) is types.NoneType:
    #     ina[index - a] = 0
    # if type(rec[index - a]) is types.NoneType:
    #     rec[index - a] = 0
    # if type(ppe[index - a]) is types.NoneType:
    #     ppe[index - a] = 0
    # if type(ta[index - a]) is types.NoneType:
    #     ta[index - a] = 0
    # if type(pay[index - a]) is types.NoneType:
    #     pay[index - a] = 0
    # if type(pat[index - a]) is types.NoneType:
    #     pat[index - a] = 0
    # ul is chan
    # ul.append((gp[index - a] + np[index - a]) / ((cb[index - a] + ina[index - a]) - (
    #           rec[index - a] + ppe[index - a] + ta[index - a] + pay[index - a] + pat[index - a])))

    # ul is return
    # ul.append(gp[index - a]/oi[index - a])

    # ws.active.cell(row=index, column=4).value = ul[index - a]

# SET ACTIVE SHEET TO SHEET 1
# ws.active = sheet1
# -----SET THE COLUMN HEADERS OF THE NEW WORKBOOK------
ws.active.cell(row=1, column=1).value = "COMPANY NAME"
ws.active.cell(row=1, column=2).value = "CHANGE IN GROSS WRITTEN PREMIUM"
ws.active.cell(row=1, column=3).value ="CHANGE IN NET WRITTEN PRERMIUM"
ws.active.cell(row=1, column=4).value ="NET INSURANCE RISK RATIO"
ws.active.cell(row=1, column=5).value ="GROSS INSURANCE RISK RATIO"
ws.active.cell(row=1, column=6).value ="CHANGE IN CAPITAL & SURPLUS"
ws.active.cell(row=1, column=7).value ="CLAIMS RESERVE RATIO"
ws.active.cell(row=1, column=8).value ="RETENTION RATIO"
ws.active.cell(row=1, column=9).value ="INVESTMENT YIELD"
ws.active.cell(row=1, column=10).value ="RETURN ON ASSETS"
ws.active.cell(row=1, column=11).value ="EXPENSE RATIO"
ws.active.cell(row=1, column=12).value ="LIQUIDITY RATIO"
ws.active.cell(row=1, column=13).value ="TECHNICAL RESERVE COVER"
ws.active.cell(row=1, column=14).value ="RETURN ON EQUITY"
ws.active.cell(row=1, column=15).value ="COMBINED RATIO"
ws.active.cell(row=1, column=16).value ="CLAIMS RATIO"
ws.active.cell(row=1, column=17).value ="PROPORTION OF INVESTMENT"


#SET ACTIVE SHEET TO SHEET 2
# ws.active = sheet2
#
# # -----SET THE COLUMN HEADERS OF THE NEW WORKSHEET------
# ws.active.cell(row=1, column=1).value = "COMPANY NAME"
# ws.active.cell(row=1, column=2).value = "Gross Claims Paid"

ws.save('D:\excel\Ratios.xlsx')
