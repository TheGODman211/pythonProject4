#
# #-----Importing all the necessary libraries----
# from openpyxl import load_workbook
# import openpyxl
# import glob
# import types
# import os,re
# #-----END OF IMPORT-----
#
# #-----SPECIFY THE PATH CONTAINING SDR SHEETS -----
#
# path = "D:\excel\Q2 2022"
#
# #---- FUNCTION TO SEARCH FOR FILES IN SPECIFIED PATH AND SAVE TO A LIST
# excel_file = glob.glob(os.path.join(path, '*.xlsx'))
#         #excel_file = glob.glob("D:\excel\Q4 2022\Life\*.xlsx", recursive=True) ------this is another to search for the files
# print(excel_file)
# new = []
#
# #-------ITERATE THROUGH SDRs AND REMOVE THE .XLSX EXTENSION TO GET THE FILE NAME-------
# for file in excel_file:
#     new.append(file[:-4])
#
# #------ISTANNTIATE EMPTY LISTS WHICH WE WILL USE LATER-------
# gp ,np,ni,gbp,nbp,me,ce,ur,ini,oi,ci,pat,cb,ina,rec,ppe,ta,tp,pay,name,ul,gl,tl,cl,nle,dp,anu,tpd,oap = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
#
# #----ITERATE THROUGH SDRs TO ALLOW US PERFORM NEEDED OPERATIONS-----
# for i,file in enumerate(excel_file):
#     # if file[-3:] == 'xls':
#     #     p.save_book_as(filename = file, dest_filename = new[i] )
#
#     wb = load_workbook(file,  data_only=True) # load the whole SDR into python and save to a variable 'wb'
#     wb.active = wb["SDR2i"]   # select the particular SDR sheet you want to work with
#     gp.append(wb.active.cell(row =26, column=3).value) # select cell value you want and add it to one of the list so can work with it
#     np.append(wb.active.cell(row =16, column=3).value)
#     # ni.append(wb.active.cell(row =14, column=2).value)     # all these lines of code have been commented out for now
#     # gbp.append(wb.active.cell(row =15, column=2).value)
#     # nbp.append(wb.active.cell(row =16, column=2).value)
#     # me.append(wb.active.cell(row =20, column=2).value)
#     # ce.append(wb.active.cell(row =38, column=2).value)
#     # ur.append(wb.active.cell(row =35, column=4).value)
#     # ini.append(wb.active.cell(row =44, column=4).value)
#     # oi.append(wb.active.cell(row =49, column=4).value)
#     # ci.append(wb.active.cell(row =33, column=4).value)
#     # pat.append(wb.active.cell(row =53, column=4).value)
#
#
#     wb.active = wb['SDR2']
#
#     cb.append(wb.active.cell(row =11, column=3).value)
#     ina.append(wb.active.cell(row=28, column=3).value)
#     rec.append(wb.active.cell(row=24, column=3).value)
#     ppe.append(wb.active.cell(row=25, column=3).value)
#     ta.append(wb.active.cell(row=20, column=3).value)
#     pay.append(wb.active.cell(row=21, column=3).value)
#     pat.append(wb.active.cell(row=22, column=3).value)
#     # #
#     wb.active = wb['SDR1']
#     name.append(wb.active.cell(row=1, column=2).value)
#     # pay.append(wb.active.cell(row=51, column=3).value)
#     #
#     # wb.active =wb['SDR8']
#     # ul.append()
#     # gl.append(wb.active.cell(row=13, column=6).value)
#     # tl.append(wb.active.cell(row=13, column=8).value)
#     # cl.append(wb.active.cell(row=13, column=10).value)
#     # nle.append(wb.active.cell(row=13, column=12).value)
#     # dp.append(wb.active.cell(row=13, column=14).value)
#     # anu.append(wb.active.cell(row=13, column=16).value)
#     # tpd.append(wb.active.cell(row=13, column=18).value)
#     # oap.append(wb.active.cell(row=13, column=20).value)
#     print(file)
# #------CREATE A NEW WORKBOOK------- and save to a variable 'ws'
# ws = openpyxl.Workbook()
# #ws = load_workbook('D:\excel\due.xlsx',  data_only=True)
# #print(gp)
#
# #print(excel_file)
# a= 2
#
# #----WRITE VALUES INTO THE NEW WORKBOOK-----
# for index, value in enumerate(gp, start=a):
#     #ws.active.cell(row=index,column=1).value = gp[index-a]
#     ws.active.cell(row=index, column=2).value = np[index-a]
#     ws.active.cell(row=index, column=3).value = cb[index - 2]
#     ws.active.cell(row=index, column=4).value = gp[index - 2]
#     ws.active.cell(row=index, column=5).value = ina[index - 2]
#     ws.active.cell(row=index, column=6).value = rec[index - 2]
#     ws.active.cell(row=index, column=7).value = ppe[index - 2]
#     ws.active.cell(row=index, column=8).value = ta[index - 2]
#     ws.active.cell(row=index, column=9).value = pay[index - 2]
#     ws.active.cell(row=index, column=10).value = pat[index - 2]
#     # ws.active.cell(row=index, column=11).value = ci[index - 2]
#     # ws.active.cell(row=index, column=12).value = pat[index - 2]
#     # ws.active.cell(row=index, column=13).value = cb[index - 2]
#     # ws.active.cell(row=index, column=14).value = ina[index - 2]
#     # ws.active.cell(row=index, column=15).value = rec[index - 2]
#     # ws.active.cell(row=index, column=16).value = ppe[index - 2]
#     # ws.active.cell(row=index, column=17).value = ta[index - 2]
#     # ws.active.cell(row=index, column=3).value = tp[index - a]
#     # ws.active.cell(row=index, column=4).value = pay[index - a]
#     ws.active.cell(row=index, column=1).value = name[index - a]
#
#     # # ws.active.cell(row=index, column=23).value = tl[index - 2]
#     # ws.active.cell(row=index, column=24).value = cl[index - 2]
#     # ws.active.cell(row=index, column=25).value = nle[index - 2]
#     # ws.active.cell(row=index, column=26).value = dp[index - 2]
#     # ws.active.cell(row=index, column=27).value = anu[index - 2]
#     # ws.active.cell(row=index, column=28).value = tpd[index - 2]
#     # ws.active.cell(row=index, column=29).value = oap[index - 2]
#
#
#
#
#
#     # print(index)
#     # print(name)
#
#
# #-----SET THE COLUMN HEADERS OF THE NEW WORKBOOK------
# ws.active.cell(row=1, column=1).value ="NAME"
# ws.active.cell(row=1, column=2).value ="TECHNICAL PROVISIONS"
# ws.active.cell(row=1, column=3).value ="CASH"
# ws.active.cell(row=1, column=4).value ="OTOTAL PAYABLES"
# ws.active.cell(row=1, column=5).value ="TOTAL INVESTMENTS"
# ws.active.cell(row=1, column=6).value ="LAND INVESTMENT"
# ws.active.cell(row=1, column=7).value ="LOANS TO PEOPLE"
# ws.active.cell(row=1, column=8).value ="LIQUIDITY RATIO"
# ws.active.cell(row=1, column=9).value ="GSE"
# ws.active.cell(row=1, column=10).value ="OTHER SECURITIES"
# ws.active.cell(row=1, column=11).value ="EQUITY MUTUAL FUNDS"
# # ws.active.cell(row=1, column=12).value ="PAT"
# # ws.active.cell(row=1, column=13).value ="Cash Balance"
# # ws.active.cell(row=1, column=14).value ="Inv Assest"
# # ws.active.cell(row=1, column=15).value ="Receivables"
# # ws.active.cell(row=1, column=16).value ="PPEs"
# # ws.active.cell(row=1, column=17).value ="Total Asset"
# # ws.active.cell(row=1, column=18).value ="Technical Provision"
# # ws.active.cell(row=1, column=19).value ="Payables"
# # ws.active.cell(row=1, column=21).value ="Universal"
# # ws.active.cell(row=1, column=22).value ="Group"
# # ws.active.cell(row=1, column=23).value ="Term"
# # ws.active.cell(row=1, column=24).value ="Credit"
# # ws.active.cell(row=1, column=25).value ="Whole Life"
# # ws.active.cell(row=1, column=26).value ="Dread disease"
# # ws.active.cell(row=1, column=27).value ="Annuities"
# # ws.active.cell(row=1, column=28).value ="Total And Permanent disability"
# # ws.active.cell(row=1, column=29).value ="Other Approved Product"
#
#
# #----SAVE THE NEW WORKBOOK
# ws.save('D:\excel\liquidityDONE.xlsx')

from openpyxl import load_workbook
import openpyxl
import glob

excel_file = glob.glob( r"D:\excel\Bew folder\2023\Q2\Non life\[!~]*.xlsx", recursive=True)

#excel_file.append(glob.glob("C:\Kojo\NonLife\.xlsx"))
print(excel_file)
new = []
for file in excel_file:
    new.append(file[:-4])
gp, fire, motor,pers, marine,liab, fin, eng,other ,np,ni,gbp,nbp,me,ce,ur,ini,oi,ci,pat,cb,ina,rec,ppe,ta,tp,pay,name = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
for i,file in enumerate(excel_file):
    # if file[-3:] == 'xls':
    #     p.save_book_as(filename = file, dest_filename = new[i] )

    wb = load_workbook(file, data_only=True)
    wb.active = wb["SDR3"]
    gp.append(wb.active.cell(row =12, column=4).value)
    np.append(wb.active.cell(row =15, column=4).value)
    ni.append(wb.active.cell(row =18, column=4).value)
    gbp.append(wb.active.cell(row =21, column=4).value)

    nbp.append(wb.active.cell(row =23, column=4).value)
    name.append(wb.active.cell(row=1, column=1).value)
    me.append(wb.active.cell(row =25, column=4).value)
    ce.append(wb.active.cell(row =24, column=4).value)
    ur.append(wb.active.cell(row =30, column=4).value)
    ini.append(wb.active.cell(row =40, column=4).value)
    oi.append(wb.active.cell(row =45, column=4).value)
    ci.append(wb.active.cell(row =28, column=4).value)
    pat.append(wb.active.cell(row =49, column=4).value)

    wb.active = wb['SDR2']

    cb.append(wb.active.cell(row =11, column=3).value)
    ina.append(wb.active.cell(row=28, column=3).value)
    rec.append(wb.active.cell(row=40, column=3).value)
    ppe.append(wb.active.cell(row=48, column=3).value)
    ta.append(wb.active.cell(row=61, column=3).value)

    wb.active = wb['SDR2i']
    tp.append(wb.active.cell(row=16, column=3).value)
    pay.append(wb.active.cell(row=26, column=3).value)

    wb.active = wb['SDR4i']
    fire.append(wb.active.cell(row=14, column=6).value)
    motor.append(wb.active.cell(row=15, column=6).value)
    pers.append(wb.active.cell(row=16, column=6).value)
    marine.append(wb.active.cell(row=17, column=6).value)
    liab.append(wb.active.cell(row=18, column=6).value)
    fin.append(wb.active.cell(row=19, column=6).value)
    eng.append(wb.active.cell(row=20, column=6).value)
    other.append(wb.active.cell(row=21, column=6).value)

ws = openpyxl.Workbook()
print(gp)

print(excel_file)
print(nbp)

a= 2
for index, value in enumerate(gp, start=a):
    ws.active.cell(row=index,column=1).value = gp[index-2]
    ws.active.cell(row=index, column=2).value = np[index-2]
    ws.active.cell(row=index, column=3).value = ni[index - 2]
    ws.active.cell(row=index, column=4).value = gbp[index - 2]
    ws.active.cell(row=index, column=5).value = nbp[index - 2]
    ws.active.cell(row=index, column=6).value = me[index - 2]
    ws.active.cell(row=index, column=7).value = ce[index - 2]
    ws.active.cell(row=index, column=8).value = ur[index - 2]
    ws.active.cell(row=index, column=9).value = ini[index - 2]
    ws.active.cell(row=index, column=10).value = oi[index - 2]
    ws.active.cell(row=index, column=11).value = ci[index - 2]
    ws.active.cell(row=index, column=12).value = pat[index - 2]
    ws.active.cell(row=index, column=13).value = cb[index - 2]
    ws.active.cell(row=index, column=14).value = ina[index - 2]
    ws.active.cell(row=index, column=15).value = rec[index - 2]
    ws.active.cell(row=index, column=16).value = ppe[index - 2]
    ws.active.cell(row=index, column=17).value = ta[index - 2]
    ws.active.cell(row=index, column=18).value = tp[index - 2]
    ws.active.cell(row=index, column=19).value = pay[index - 2]
    ws.active.cell(row=index, column=20).value = name[index - 2]
    ws.active.cell(row=index, column=21).value = motor[index - 2]
    ws.active.cell(row=index, column=22).value = fire[index - 2]
    ws.active.cell(row=index, column=23).value = pers[index - 2]
    ws.active.cell(row=index, column=24).value = marine[index - 2]
    ws.active.cell(row=index, column=25).value = liab[index - 2]
    ws.active.cell(row=index, column=26).value = fin[index - 2]
    ws.active.cell(row=index, column=27).value = eng[index - 2]
    ws.active.cell(row=index, column=28).value = other[index - 2]



    print(index)
ws.active.cell(row=1, column=1).value ="Gross premium"
ws.active.cell(row=1, column=2).value ="Net premium"
ws.active.cell(row=1, column=3).value ="Net income"
ws.active.cell(row=1, column=4).value ="Gross Claims Incurred"
ws.active.cell(row=1, column=5).value ="Net Claims Incurred"
ws.active.cell(row=1, column=6).value ="mgt expense"
ws.active.cell(row=1, column=7).value ="Comission expense"
ws.active.cell(row=1, column=9).value ="Underwriting Results"
ws.active.cell(row=1, column=10).value ="Investment Income"
ws.active.cell(row=1, column=11).value ="Other Income"
ws.active.cell(row=1, column=8).value ="Commission Income"
ws.active.cell(row=1, column=12).value ="PAT"
ws.active.cell(row=1, column=13).value ="Cash Balance"
ws.active.cell(row=1, column=14).value ="Inv Assest"
ws.active.cell(row=1, column=15).value ="Receivables"
ws.active.cell(row=1, column=16).value ="PPEs"
ws.active.cell(row=1, column=17).value ="Total Asset"
ws.active.cell(row=1, column=18).value ="Technical Provision"
ws.active.cell(row=1, column=19).value ="Payables"
ws.active.cell(row=1, column=20).value = 'Name'
ws.active.cell(row=1, column=21).value ="Fire"
ws.active.cell(row=1, column=22).value ="Motor"
ws.active.cell(row=1, column=23).value ="Personal Accident"
ws.active.cell(row=1, column=24).value ="Marine"
ws.active.cell(row=1, column=25).value ="Liability"
ws.active.cell(row=1, column=26).value ="Fin guarantee"
ws.active.cell(row=1, column=27).value ="Engineering"
ws.active.cell(row=1, column=28).value = 'Other'


ws.save(r'D:\excel\q2 market nl.xlsx')