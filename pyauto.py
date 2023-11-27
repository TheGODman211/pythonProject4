import os
import openpyxl
import time
import pyautogui


# ... [Your existing code above this] ...
wb = openpyxl.load_workbook(r"D:\excel\Annual.xlsx", data_only=True)

    # Open the first sheet
sheet = wb.worksheets[0]

    # Iterate through the rows in the specified column and append values to a list
values_list = []
values_list1 = []
values_list2 = []
values_list3 = []
values_list4 = []
values_list5 = []
for row in sheet.iter_rows(min_col=6, max_col=6):
    for cell in row:
            values_list.append(cell.value)

for row in sheet.iter_rows(min_col=2, max_col=2):
    for cell in row:
            values_list1.append(cell.value)

for row in sheet.iter_rows(min_col=3, max_col=3):
    for cell in row:
            values_list2.append(cell.value)

for row in sheet.iter_rows(min_col=4, max_col=4):
    for cell in row:
            values_list3.append(cell.value)

for row in sheet.iter_rows(min_col=5, max_col=5):
    for cell in row:
            values_list4.append(cell.value)

print(values_list3)
print(len(values_list3))


def update_value_with_keyboard(filepath, value_c11, value_d28):
    os.startfile(filepath)
    time.sleep(5)  # Wait for a few seconds to allow Excel to open

    # Navigate to C11 and update
    pyautogui.hotkey('ctrl', 'g')  # 'ctrl+g' opens the 'Go To' window in Excel
    pyautogui.write('SDR2!C11')
    pyautogui.press('enter')
    pyautogui.write(str(value_c11))
    pyautogui.press('enter')

    time.sleep(2)

    # Navigate to D28 and update
    pyautogui.hotkey('ctrl', 'g')
    pyautogui.write('SDR3!D28')  # Assuming the 7th sheet is named "Sheet7". Modify accordingly.
    pyautogui.press('enter')
    time.sleep(3)
    pyautogui.write(str(value_d28))
    pyautogui.press('enter')

    # Save and close
    pyautogui.hotkey('ctrl', 's')  # Save
    time.sleep(1)  # Wait for save to complete
    pyautogui.hotkey('alt', 'f4')  # Close Excel

    time.sleep(3)


def update_value_in_workbooks(directory_path):
    # List all files in the directory
    files = [f for f in os.listdir(directory_path) if os.path.isfile(os.path.join(directory_path, f))]
    print(len(files))
    for i, file in enumerate(files):
        # Check if the file is an Excel workbook
        if file.endswith('.xlsx') or file.endswith('.xlsm'):
            filepath = os.path.join(directory_path, file)

            # Load the workbook
            workbook = openpyxl.load_workbook(filepath, data_only=True)

            # Determine the new values for C11 and D28
            sheet = workbook.worksheets[2]
            new_value_c11 = int(sheet['C11'].value) + values_list[i + 1] if sheet['C11'].value else values_list[i + 1]

            sheet1 = workbook.worksheets[6]
            new_value_d28 = int(sheet1['D28'].value) + values_list[i + 1] if sheet1['D28'].value else values_list[
                i + 1]

            update_value_with_keyboard(filepath, new_value_c11, new_value_d28)
            print(f"Updated {file}")


# Use the function
directory_path = r"D:\excel\2018 Annual SDR\NON LIFE COMPANIES"
update_value_in_workbooks(directory_path)




# import win32com.client as win32
# import os
# import openpyxl
#
# # ... [Your code to extract values_list1 from Annual.xlsx] ...
# wb = openpyxl.load_workbook(r"D:\excel\Annual.xlsx", data_only=True)
#
#     # Open the first sheet
# sheet = wb.worksheets[0]
#
#     # Iterate through the rows in the specified column and append values to a list
# values_list = []
# values_list1 = []
# values_list2 = []
# values_list3 = []
# values_list4 = []
# values_list5 = []
# for row in sheet.iter_rows(min_col=6, max_col=6):
#     for cell in row:
#             values_list.append(cell.value)
#
# for row in sheet.iter_rows(min_col=2, max_col=2):
#     for cell in row:
#             values_list1.append(cell.value)
#
# for row in sheet.iter_rows(min_col=3, max_col=3):
#     for cell in row:
#             values_list2.append(cell.value)
#
# for row in sheet.iter_rows(min_col=4, max_col=4):
#     for cell in row:
#             values_list3.append(cell.value)
#
# for row in sheet.iter_rows(min_col=5, max_col=5):
#     for cell in row:
#             values_list4.append(cell.value)
#
# print(values_list1)
# print(len(values_list1))
#
#
#
# def update_value_in_workbooks(directory_path):
#     # List all files in the directory
#     files = [f for f in os.listdir(directory_path) if os.path.isfile(os.path.join(directory_path, f))]
#     print(len(files))
#
#     # Start an instance of Excel
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     excel.Visible = False  # Run Excel in the background
#
#     for i, file in enumerate(files):
#         # Check if the file is an Excel workbook
#         if file.endswith('.xlsx') or file.endswith('.xlsm'):
#             filepath = os.path.join(directory_path, file)
#
#             # Load the workbook
#             wb = excel.Workbooks.Open(filepath)
#
#             # Enable Editing (if needed)
#             wb.CheckInWithVersion(True)
#
#             # Access the sheets
#             sheet = wb.Sheets[3]
#             sheet1 = wb.Sheets[7]
#
#             # Update the value in cell C11 in the first sheet
#             if isinstance(sheet.Range("C11").Value, (int, float)):
#                 sheet.Range("C11").Value += values_list1[i + 1]
#             else:
#                 sheet.Range("C11").Value = values_list1[i + 1]
#
#             # Update the value in cell D28 in the second sheet
#             if isinstance(sheet1.Range("D28").Value, (int, float)):
#                 sheet1.Range("D28").Value += values_list1[i + 1]
#             else:
#                 sheet1.Range("D28").Value = values_list1[i + 1]
#
#             # Save and close the workbook
#             wb.Save()
#             wb.Close()
#             print(f"Updated {file}")
#
#     # Quit Excel
#     excel.Quit()
#
#
# # Use the function
# directory_path = r"D:\excel\2018 Annual SDR\NON-LIFE 2018"
# update_value_in_workbooks(directory_path)
