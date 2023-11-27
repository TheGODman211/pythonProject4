import os
import openpyxl

wb = openpyxl.load_workbook(r"D:\excel\Annual.xlsx", data_only=True)

    # Open the first sheet
sheet = wb.worksheets[0]

    # Iterate through the rows in the specified column and append values to a list
values_list = []
values_list1 = []
values_list2 = []
values_list3 = []
values_list4 = []
values_list5 = []
for row in sheet.iter_rows(min_col=6, max_col=6):
    for cell in row:
            values_list.append(cell.value)

for row in sheet.iter_rows(min_col=2, max_col=2):
    for cell in row:
            values_list1.append(cell.value)

for row in sheet.iter_rows(min_col=3, max_col=3):
    for cell in row:
            values_list2.append(cell.value)

for row in sheet.iter_rows(min_col=4, max_col=4):
    for cell in row:
            values_list3.append(cell.value)

for row in sheet.iter_rows(min_col=5, max_col=5):
    for cell in row:
            values_list4.append(cell.value)

print(values_list1)
print(len(values_list1))

def update_value_in_workbooks(directory_path):


    # List all files in the directory
    files = [f for f in os.listdir(directory_path) if os.path.isfile(os.path.join(directory_path, f))]
    print(files)
    for i, file in enumerate(files):
        # Check if the file is an Excel workbook
        if file.endswith('.xlsx') or file.endswith('.xlsm'):
            filepath = os.path.join(directory_path, file)
            print(file)
            # Load the workbook
            workbook = openpyxl.load_workbook(filepath,data_only=True)

            # Open the first sheet
            sheet = workbook.worksheets[2]

            # Update the value in cell C11
            if sheet['C11'].value:
                print(sheet['C11'].value)
                print(values_list1[i+1])
                sheet['C11'] = int(sheet['C11'].value) + values_list1[i+1]
            else:
                sheet['C11'] = values_list1[i+1]

            sheet1 = workbook.worksheets[6]
            if sheet1['D28'].value:
                print(sheet1['D28'].value)
                print(values_list1[i+1])
                sheet1['D28'] = int(sheet['D28'].value) + values_list1[i+1]
            else:
                sheet1['D28'] = values_list1[i+1]

            # Save the workbook
            workbook.save(filepath)
            print(f"Updated {file}")


# Use the function
directory_path = r"D:\excel\2018 Annual SDR\NON-LIFE 2018"
update_value_in_workbooks(directory_path)


# import openpyxl
#
#
# def get_values_from_column(workbook_path):
#     # Load the workbook
#     wb = openpyxl.load_workbook(workbook_path)
#
#     # Open the first sheet
#     sheet = wb.worksheets[0]
#
#     # Iterate through the rows in the specified column and append values to a list
#     values_list = []
#     values_list1 = []
#     values_list2 = []
#     values_list3 = []
#     values_list4 = []
#     values_list5 = []
#     for row in sheet.iter_rows(min_col=6, max_col=6):
#         for cell in row:
#             values_list.append(cell.value)
#
#     for row in sheet.iter_rows(min_col=2, max_col=2):
#         for cell in row:
#             values_list1.append(cell.value)
#
#     for row in sheet.iter_rows(min_col=3, max_col=3):
#         for cell in row:
#             values_list2.append(cell.value)
#
#     for row in sheet.iter_rows(min_col=4, max_col=4):
#         for cell in row:
#             values_list3.append(cell.value)
#
#     for row in sheet.iter_rows(min_col=5, max_col=5):
#         for cell in row:
#             values_list4.append(cell.value)
#
#     return values_list
#



# # Use the function
# workbook_path = r"D:\excel\Annual.xlsx"  # Replace with your workbook path
# column_values = get_values_from_column(workbook_path)
# print(column_values)
