import os
import openpyxl
import datetime

# Path to the folder containing the files
folder_path = r'D:\excel\Sign in\completed'

# Create a new workbook to store the results
result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active

# Iterate through files in the folder
for filename in os.listdir(folder_path):
    print(filename)
    file_path = os.path.join(folder_path, filename)

    # Check if the file is an Excel workbook
    if file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_obj = workbook.active

        date= []
        time_list = []
        name = []
        date1 = []
        time_list1 = []
        name1 = []
        date2 = []
        time_list2 = []
        name2 = []
        date3 = []
        time_list3 = []
        name3 = []
        row_num = 1
        i = 0

        while i < 4:

            for idx, times in enumerate(sheet_obj.iter_rows(min_row=row_num,max_row=row_num+31, min_col=1, max_col=12,  values_only=True),start = row_num+3 ,):
                # date.append(sheet_obj.cell(row=row_num, column=2).value)
                print(sheet_obj.cell(row=idx, column=2))
                print(type(sheet_obj.cell(row=idx, column=2).value))
                if isinstance(sheet_obj.cell(row=idx, column=2).value, datetime.time) and sheet_obj.cell(row=idx, column=2).value > datetime.time(8, 0, 0):
                    name.append(sheet_obj.cell(row=idx, column=1).value)
                    time_list.append(sheet_obj.cell(row=idx, column=2).value)
                    date.append(sheet_obj.cell(row=row_num, column=2).value)
                    print(sheet_obj.cell(row=row_num, column=2))
                    print(name)
                    print(time_list)
                    print(date)
                if isinstance(sheet_obj.cell(row=idx, column=4).value, datetime.time) and sheet_obj.cell(row=idx,
                                                                                                         column=4).value > datetime.time(
                        8, 0, 0):
                    name1.append(sheet_obj.cell(row=idx, column=1).value)
                    time_list1.append(sheet_obj.cell(row=idx, column=4).value)
                    date1.append(sheet_obj.cell(row=row_num, column=4).value)
                if isinstance(sheet_obj.cell(row=idx, column=6).value, datetime.time) and sheet_obj.cell(row=idx,
                                                                                                         column=6).value > datetime.time(
                        8, 0, 0):
                    name2.append(sheet_obj.cell(row=idx, column=1).value)
                    time_list2.append(sheet_obj.cell(row=idx, column=6).value)
                    date2.append(sheet_obj.cell(row=row_num, column=6).value)
                if isinstance(sheet_obj.cell(row=idx, column=8).value, datetime.time) and sheet_obj.cell(row=idx,
                                                                                                         column=8).value > datetime.time(
                        8, 0, 0):
                    name3.append(sheet_obj.cell(row=idx, column=1).value)
                    time_list3.append(sheet_obj.cell(row=idx, column=8).value)
                    date3.append(sheet_obj.cell(row=row_num, column=8).value)
            row_num+=36
            i+=1

    wb = openpyxl.Workbook()
    longest_list =[name,name1, name2, name3]

    longest_length = max(len(name), len(name1), len(name2), len(name3))

    for i in range(longest_length):
        if i < len(name):
            wb.active.cell(row=i + 1, column=1).value = name[i]
        if i < len(date):
            wb.active.cell(row=i + 1, column=2).value = date[i]
        if i < len(time_list):
            wb.active.cell(row=i + 1, column=3).value = time_list[i]
        if i < len(name1):
            wb.active.cell(row=i + 1, column=4).value = name1[i]
        if i < len(date1):
            wb.active.cell(row=i + 1, column=5).value = date1[i]
        if i < len(time_list1):
            wb.active.cell(row=i + 1, column=6).value = time_list1[i]
        if i < len(name2):
            wb.active.cell(row=i + 1, column=7).value = name2[i]
        if i < len(date2):
            wb.active.cell(row=i + 1, column=8).value = date2[i]
        if i < len(time_list2):
            wb.active.cell(row=i + 1, column=9).value = time_list2[i]
        if i < len(name3):
            wb.active.cell(row=i + 1, column=10).value = name3[i]
        if i < len(date3):
            wb.active.cell(row=i + 1, column=11).value = date3[i]
        if i < len(time_list3):
            wb.active.cell(row=i + 1, column=12).value = time_list3[i]

    # for i, val in enumerate(max(longest_list, key=len), start=1):
    #     wb.active.cell(row=i, column = 1).value = name[i-1]
    #     wb.active.cell(row=i, column=2).value = date[i - 1]
    #     wb.active.cell(row=i, column=3).value = time_list[i - 1]
    #     wb.active.cell(row=i, column=4).value = name1[i - 1]
    #     wb.active.cell(row=i, column=5).value = date1[i - 1]
    #     wb.active.cell(row=i, column=6).value = time_list1[i - 1]
    #     wb.active.cell(row=i, column=7).value = name2[i - 1]
    #     wb.active.cell(row=i, column=8).value = date2[i - 1]
    #     wb.active.cell(row=i, column=9).value = time_list2[i - 1]
    #     wb.active.cell(row=i, column=10).value = name3[i - 1]
    #     wb.active.cell(row=i, column=11).value = date3[i - 1]
    #     wb.active.cell(row=i, column=12).value = time_list3[i - 1]
        # wb.active.cell(row=i, column=1).value = name[i - 1]
        # wb.active.cell(row=i, column=1).value = name[i - 1]

    wb.save(os.path.join(folder_path, filename[:-4] + "late.xlsx"))



    # break

        #
        #         # i+=1



